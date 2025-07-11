from flask import Flask, render_template_string, request, redirect
from openpyxl import load_workbook
from datetime import datetime

app = Flask(__name__)

# Caminho do seu arquivo Excel
ARQUIVO_EXCEL = 'Inventario_Impressoras.xlsx'
NOME_PLANILHA = 'Sheet1'

# Página HTML simples
html = """
<!DOCTYPE html>
<html> 
<head>
    <title>Registro de Toner</title>
</head>
<body>
    <h2>Registrar Troca de Toner</h2>
    <form method="POST">
        Impressora: <input type="text" name="impressora" required><br><br>
        Toner: <input type="text" name="toner" required><br><br>
        Responsável: <input type="text" name="responsavel"><br><br>
        <button type="submit">Registrar</button>
    </form>
    <p>{{ mensagem }}</p>
</body>
</html>
"""

@app.route("/", methods=["GET", "POST"])
def registrar():
    mensagem = ""
    if request.method == "POST":
        impressora = request.form["impressora"]
        toner = request.form["toner"]
        responsavel = request.form["responsavel"]
        data_hoje = datetime.now().strftime("%d/%m/%Y %H:%M")

        try:
            # Abrir e editar a planilha
            wb = load_workbook(ARQUIVO_EXCEL)
            sheet = wb[NOME_PLANILHA]

            # Inserir na próxima linha vazia
            proxima_linha = sheet.max_row + 1
            sheet.cell(row=proxima_linha, column=1).value = data_hoje
            sheet.cell(row=proxima_linha, column=2).value = impressora
            sheet.cell(row=proxima_linha, column=3).value = toner
            sheet.cell(row=proxima_linha, column=4).value = responsavel

            wb.save(ARQUIVO_EXCEL)

            mensagem = "Registro salvo com sucesso!"

        except Exception as e:
            mensagem = f"Erro ao salvar: {e}"

    return render_template_string(html, mensagem=mensagem)

if __name__ == "__main__":
    app.run(debug=True)
